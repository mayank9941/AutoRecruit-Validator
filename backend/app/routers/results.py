"""
Results endpoint -- shows candidates grouped/filtered by their final
screening status (Eligible / Not Eligible / Needs Review / Not Evaluated),
plus an Excel export sorted with Eligible candidates first.
"""

import io
import re
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session, joinedload

from app.db.session import get_db
from app.models.job_profile import JobProfile
from app.models.candidate import Candidate
from app.models.criterion_evaluation import CriterionEvaluation
from app.schemas.results import ResultsResponse, ResultsSummary
from app.services.auth_service import get_current_hr_user

router = APIRouter(
    prefix="/jd/profiles/{profile_id}/results",
    tags=["results"],
    dependencies=[Depends(get_current_hr_user)],
)

# Eligible candidates should always appear first, followed by Needs Review,
# then Not Eligible, then anyone not yet evaluated.
STATUS_SORT_ORDER = {
    "eligible": 0,
    "needs_review": 1,
    "not_eligible": 2,
    "not_evaluated": 3,
}


def _get_profile_or_404(profile_id: str, db: Session) -> JobProfile:
    profile = db.query(JobProfile).filter(JobProfile.id == profile_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail=f"Job Profile '{profile_id}' not found")
    return profile


def _build_summary(all_candidates: list[Candidate]) -> ResultsSummary:
    return ResultsSummary(
        total=len(all_candidates),
        eligible=sum(1 for c in all_candidates if c.status == "eligible"),
        not_eligible=sum(1 for c in all_candidates if c.status == "not_eligible"),
        needs_review=sum(1 for c in all_candidates if c.status == "needs_review"),
        not_evaluated=sum(1 for c in all_candidates if c.status == "not_evaluated"),
    )


@router.get("", response_model=ResultsResponse)
def get_results(
    profile_id: str,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Returns a summary of candidate counts by status, plus the candidate
    list sorted with Eligible first. Pass ?status=eligible (or
    not_eligible / needs_review / not_evaluated) to filter to just one
    bucket.
    """
    _get_profile_or_404(profile_id, db)

    all_candidates = db.query(Candidate).filter(Candidate.job_profile_id == profile_id).all()
    summary = _build_summary(all_candidates)

    filtered = all_candidates if status is None else [c for c in all_candidates if c.status == status]
    sorted_candidates = sorted(filtered, key=lambda c: STATUS_SORT_ORDER.get(c.status, 99))

    return ResultsResponse(summary=summary, candidates=sorted_candidates)


# ---------------------------------------------------------------------------
# Excel export -- mirrors IHMCL's manual "Screening Sheet" format: a grouped
# two-row header (Essential Qualification / Desirable Qualification /
# Essential Experience / Preferred Experience sections, each with
# possessed / fulfilled / documents-attached sub-columns), one row per
# candidate, Eligible candidates first.
# ---------------------------------------------------------------------------

_YES_NO = {True: "Yes", False: "No"}


def _parse_dob(dob: Optional[str]) -> Optional[date]:
    if not dob:
        return None
    text = dob.strip()[:19]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _age_from_dob(dob: Optional[str]) -> Optional[int]:
    parsed = _parse_dob(dob)
    if parsed is None:
        return None
    today = date.today()
    return today.year - parsed.year - ((today.month, today.day) < (parsed.month, parsed.day))


def _excel_field(candidate: Candidate, *keys: str) -> str:
    """Joins non-empty roster columns (e.g. course name + specialization)."""
    data = candidate.raw_excel_data or {}
    parts = []
    for key in keys:
        value = data.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() not in ("nan", "n/a", "na", "none"):
            parts.append(text)
    return " - ".join(parts)


def _fulfilled(evals: list[CriterionEvaluation]) -> str:
    """Collapses a group of criterion results into Yes / No / Needs review."""
    if not evals:
        return "—"
    if any(e.result == "fail" for e in evals):
        return "No"
    if any(e.result == "needs_review" for e in evals):
        return "Needs review"
    return "Yes"


def _docs_attached(evals: list[CriterionEvaluation]) -> str:
    if not evals:
        return "—"
    return _YES_NO[all(e.citation_document for e in evals)]


def _total_exp_text(evals: list[CriterionEvaluation]) -> str:
    """Pulls the computed total ("total = 7.7 yrs") out of the evaluation
    reasoning when present; falls back to the match percentage."""
    for e in evals:
        # The final total is stated last in the reasoning ("... total = 7.7
        # yrs >= 5 required"), after the per-organization sub-totals -- so
        # take the LAST match, not the first.
        matches = re.findall(r"total[^=\n]*=\s*~?([\d.]+)\s*(?:yrs|years)", e.reasoning or "", re.IGNORECASE)
        if matches:
            return f"~{matches[-1]} yrs"
    for e in evals:
        if e.match_percentage is not None:
            return f"{e.match_percentage}% of required"
    return "—"


@router.get("/export")
def export_results(
    profile_id: str,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Exports candidates as an IHMCL-style Screening Sheet, sorted with
    Eligible first. Pass ?status=... to export just one bucket.
    """
    profile = _get_profile_or_404(profile_id, db)

    all_candidates = db.query(Candidate).filter(Candidate.job_profile_id == profile_id).all()
    filtered = all_candidates if status is None else [c for c in all_candidates if c.status == status]
    sorted_candidates = sorted(filtered, key=lambda c: STATUS_SORT_ORDER.get(c.status, 99))

    evaluations = (
        db.query(CriterionEvaluation)
        .options(joinedload(CriterionEvaluation.criterion))
        .filter(CriterionEvaluation.candidate_id.in_([c.id for c in sorted_candidates] or [""]))
        .all()
    )
    evals_by_candidate: dict[str, list[CriterionEvaluation]] = {}
    for e in evaluations:
        evals_by_candidate.setdefault(e.candidate_id, []).append(e)

    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Sheet"

    # -- Header (rows 1-2) --------------------------------------------------
    # (col, group title or None, sub header, width)
    single_columns = [
        ("A", "S No.", 6),
        ("B", "Application No.", 20),
        ("C", "Applicant Name", 24),
        ("D", "Date of Birth", 14),
        ("E", "Age", 7),
        ("F", "Whether supporting documents attached (Yes/No)", 16),
        ("G", "Category", 10),
        ("H", "Graduation", 26),
        ("I", "Post Graduation", 26),
    ]
    groups = [
        ("Essential Qualification", [
            ("J", "Qualification possessed", 28),
            ("K", "Whether Fulfilled (Yes/No)", 14),
            ("L", "Whether supporting documents attached (Yes/No)", 16),
        ]),
        ("Desirable Qualification", [
            ("M", "Qualification possessed", 28),
            ("N", "Whether Fulfilled (Yes/No)", 14),
            ("O", "Whether supporting documents attached (Yes/No)", 16),
        ]),
        ("Essential Experience", [
            ("P", "Total Exp.", 14),
            ("Q", "Relevant Experience", 18),
            ("R", "Experience Fulfilled (Yes/No)", 14),
            ("S", "Whether supporting documents attached (Yes/No)", 16),
        ]),
        ("Preferred Experience / Skills", [
            ("T", "Skills & experience possessed", 40),
            ("U", "Match %", 10),
            ("V", "Gist", 50),
        ]),
    ]
    tail_columns = [
        ("W", "Eligible/Not Eligible", 18),
        ("X", "Document Remarks", 30),
        ("Y", "Remarks", 40),
    ]

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _header_cell(coordinate: str, value: str):
        cell = ws[coordinate]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for col, title, width in single_columns + tail_columns:
        ws.merge_cells(f"{col}1:{col}2")
        _header_cell(f"{col}1", title)
        ws[f"{col}2"].border = border
        ws.column_dimensions[col].width = width

    for group_title, subs in groups:
        first, last = subs[0][0], subs[-1][0]
        ws.merge_cells(f"{first}1:{last}1")
        _header_cell(f"{first}1", group_title)
        for col, sub_title, width in subs:
            ws[f"{col}1"].border = border
            _header_cell(f"{col}2", sub_title)
            ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "A3"

    # -- Data rows ------------------------------------------------------------
    status_text = {
        "eligible": "Eligible",
        "not_eligible": "Not Eligible",
        "needs_review": "Needs Review",
        "not_evaluated": "Not Screened",
    }
    docs_text = {
        "documents_complete": "Yes",
        "documents_incomplete": "Partial",
    }

    for idx, c in enumerate(sorted_candidates, start=1):
        evals = evals_by_candidate.get(c.id, [])
        edu_essential = [e for e in evals if e.criterion.type == "education" and e.criterion.is_essential]
        edu_desirable = [e for e in evals if e.criterion.type == "education" and not e.criterion.is_essential]
        exp_essential = [e for e in evals if e.criterion.type == "experience" and e.criterion.is_essential]
        skill_evals = [e for e in evals if e.criterion.type == "skill"]

        graduation = _excel_field(c, "G Course Name", "G Specialization")
        post_grad = _excel_field(c, "P Course Name", "P Specialization")
        other_qual = _excel_field(c, "O Course Name")
        qualifications = "; ".join(p for p in (graduation, post_grad, other_qual) if p)

        exp_pct = next((e.match_percentage for e in exp_essential if e.match_percentage is not None), None)
        skill_pct = next((e.match_percentage for e in skill_evals if e.match_percentage is not None), None)
        exp_gist = next((e.reasoning for e in exp_essential if e.reasoning), "")

        failed = sorted({e.criterion.type for e in evals if e.result == "fail" and e.criterion.is_essential})
        review = sorted({e.criterion.type for e in evals if e.result == "needs_review" and e.criterion.is_essential})
        if c.status == "eligible":
            remarks = "Meets all essential criteria"
        elif c.status == "not_eligible":
            remarks = f"Not eligible — fails essential: {', '.join(failed)}" if failed else "Not eligible"
        elif c.status == "needs_review":
            remarks = f"Manual review needed: {', '.join(review)}" if review else "Manual review needed"
        else:
            remarks = "Screening not run yet"
        if c.status_overridden:
            remarks += f" (status set manually by HR{': ' + c.override_reason if c.override_reason else ''})"

        doc_remarks = []
        if c.ingestion_status != "documents_complete":
            doc_remarks.append(f"Documents: {c.ingestion_status.replace('_', ' ')}")
        uncited = sorted({
            e.criterion.type
            for e in evals
            if not e.citation_document and e.criterion.type not in ("age", "skill")
        })
        if uncited:
            doc_remarks.append(f"No supporting document cited for: {', '.join(uncited)}")

        row = [
            idx,
            c.external_id,
            c.name or "",
            c.dob or "",
            _age_from_dob(c.dob) or _excel_field(c, "Age") or "",
            docs_text.get(c.ingestion_status, "No"),
            c.normalized_category or c.raw_category or "",
            graduation,
            post_grad,
            qualifications,
            _fulfilled(edu_essential),
            _docs_attached(edu_essential),
            post_grad or other_qual,
            _fulfilled(edu_desirable),
            _docs_attached(edu_desirable),
            _total_exp_text(exp_essential),
            f"{exp_pct}% of required" if exp_pct is not None else "—",
            _fulfilled(exp_essential),
            _docs_attached(exp_essential),
            next((e.reasoning for e in skill_evals if e.reasoning), ""),
            f"{skill_pct}%" if skill_pct is not None else "—",
            exp_gist,
            status_text.get(c.status, c.status),
            "; ".join(doc_remarks),
            remarks,
        ]
        row_num = idx + 2
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = body_align
            cell.border = border
            cell.font = Font(size=10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = "".join(ch if ch.isalnum() else "_" for ch in profile.title)[:50]
    filename = f"Screening_Sheet_{safe_title}_{date.today().strftime('%d.%m.%Y')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
